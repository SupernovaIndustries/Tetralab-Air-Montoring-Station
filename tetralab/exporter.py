"""Export Excel multi-sheet con grafici embedded.

Genera un .xlsx contenente:
  - Sheet "Live (1 min)"
  - Sheet "Hourly"
  - Sheet "12h"
  - Sheet "Daily"
  - Sheet "Charts" con grafici di tutte le metriche su base oraria
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import Settings
from .storage import METRICS, Storage


METRIC_LABELS = {
    "pm1":  ("PM 1.0",  "µg/m³"),
    "pm25": ("PM 2.5",  "µg/m³"),
    "pm4":  ("PM 4.0",  "µg/m³"),
    "pm10": ("PM 10",   "µg/m³"),
    "rh":   ("Umidità", "%"),
    "temp": ("Temperatura", "°C"),
    "voc":  ("VOC Index", ""),
    "nox":  ("NOx Index", ""),
}


HEADER_FILL = PatternFill(start_color="1E2130", end_color="1E2130", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(wb: Workbook, name: str, rows: list[dict], tz: ZoneInfo) -> None:
    ws = wb.create_sheet(title=name)
    headers = ["Timestamp (locale)", "Unix UTC"] + [
        f"{lbl} ({u})" if u else lbl for lbl, u in (METRIC_LABELS[m] for m in METRICS)
    ] + ["N samples"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ts = r.get("ts") or 0
        local = datetime.fromtimestamp(ts, tz=tz)
        row = [local.replace(tzinfo=None), int(ts)]
        for m in METRICS:
            row.append(r.get(m))
        row.append(int(r.get("n_samples") or 0))
        ws.append(row)

    # Larghezza colonne
    widths = [22, 12] + [14] * len(METRICS) + [10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Format colonna A come data
    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm:ss"

    ws.freeze_panes = "A2"


def _add_charts_sheet(wb: Workbook, hourly_rows: list[dict], tz: ZoneInfo) -> None:
    """Sheet con un grafico per ogni metrica usando i dati hourly."""
    if not hourly_rows:
        return
    ws = wb.create_sheet(title="Charts")
    # Scrivi tabella nascosta a destra (col Z+) per le serie
    headers = ["Timestamp"] + [METRIC_LABELS[m][0] for m in METRICS]
    ws.append(headers)
    for r in hourly_rows:
        ts = r.get("ts") or 0
        local = datetime.fromtimestamp(ts, tz=tz).replace(tzinfo=None)
        row = [local] + [r.get(m) for m in METRICS]
        ws.append(row)

    n_data_rows = len(hourly_rows)
    n_metrics = len(METRICS)

    # Crea un chart per ciascuna metrica
    for i, m in enumerate(METRICS):
        chart = LineChart()
        chart.title = f"{METRIC_LABELS[m][0]} (orario)"
        chart.style = 2
        chart.y_axis.title = METRIC_LABELS[m][1] or "valore"
        chart.x_axis.title = "ora locale"
        chart.height = 8
        chart.width = 18

        col = i + 2  # +2 perche' colonna 1 e' timestamp
        data = Reference(ws, min_col=col, min_row=1, max_row=n_data_rows + 1, max_col=col)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_data_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Posizione grafici: 2 colonne, righe ogni 18
        anchor_row = 2 + (i // 2) * 18
        anchor_col = "K" if (i % 2) == 0 else "U"
        ws.add_chart(chart, f"{anchor_col}{anchor_row}")


def build_xlsx(storage: Storage, settings: Settings,
               ts_from: Optional[int] = None, ts_to: Optional[int] = None) -> bytes:
    """Costruisce il file Excel in memoria e ritorna i bytes."""
    tz = ZoneInfo(settings.timezone_name)

    wb = Workbook()
    # Rimuovi sheet di default
    default = wb.active
    wb.remove(default)

    minute_rows = storage.fetch("minute", ts_from=ts_from, ts_to=ts_to)
    hour_rows   = storage.fetch("hour",   ts_from=ts_from, ts_to=ts_to)
    half_rows   = storage.fetch("half",   ts_from=ts_from, ts_to=ts_to)
    day_rows    = storage.fetch("day",    ts_from=ts_from, ts_to=ts_to)

    _write_sheet(wb, "Live (1 min)", minute_rows, tz)
    _write_sheet(wb, "Hourly", hour_rows, tz)
    _write_sheet(wb, "12h",    half_rows, tz)
    _write_sheet(wb, "Daily",  day_rows,  tz)
    _add_charts_sheet(wb, hour_rows, tz)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
